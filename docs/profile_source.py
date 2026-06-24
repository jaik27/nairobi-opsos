#!/usr/bin/env python3
"""
profile_source.py — Nairobi OpsOS source profiler.

The repeatable "profile any client's file" step from the Segment Onboarding
Playbook. Works on any .xlsx / .xlsm / .csv export a prospect hands you, and
prints a structured profile: sheets, detected header, completeness, controlled
vocabularies, whitespace/error anomalies, duplicate clusters, and a guess at the
source SHAPE (which adapter handles it).

Read-only. Never modifies the input.

Usage:
    python profile_source.py CLIENT_FILE.xlsx
    python profile_source.py CLIENT_FILE.xlsx --sheet "ITEM DB"
    python profile_source.py CLIENT_FILE.xlsx --key "SHORT ITEM DESCRIPTION"
    python profile_source.py data.csv

Deps: openpyxl (xlsx/xlsm), pandas optional (csv). pip install --break-system-packages openpyxl
"""
import sys, re, argparse, os
from collections import Counter, defaultdict

ERR = re.compile(r'#(REF|VALUE|N/A|DIV/0|NAME|NULL|NUM)!?', re.I)

def norm(s):
    s = re.sub(r'\s+', ' ', str(s)).strip().upper()
    s = re.sub(r'[^A-Z0-9]+', ' ', s).strip()
    return s

def load_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
        out[name] = rows
    wb.close()
    return out

def load_csv(path):
    import csv
    with open(path, newline='', encoding='utf-8', errors='replace') as f:
        return {"(csv)": [row for row in csv.reader(f)]}

def detect_header(rows, scan=25):
    """Pick the row in the first `scan` rows with the most distinct non-empty
    short string cells — the most likely real header in a messy sheet."""
    best_i, best_score = 0, -1
    for i, r in enumerate(rows[:scan]):
        cells = [str(c).strip() for c in r if c is not None and str(c).strip() != ""]
        shortish = [c for c in cells if len(c) <= 40]
        score = len(set(shortish))
        if score > best_score:
            best_i, best_score = i, score
    return best_i

def profile_sheet(name, rows, key=None):
    print(f"\n{'='*72}\nSHEET: {name}")
    rows = [r for r in rows if r is not None]
    if not rows:
        print("  (empty)"); return
    hi = detect_header(rows)
    hdr = [str(h).strip() if h not in (None, "") else f"col{j}" for j, h in enumerate(rows[hi])]
    data = [r for r in rows[hi+1:] if any(v is not None and str(v).strip() != "" for v in r)]
    N = len(data)
    print(f"  header row: {hi+1} | data rows: {N} | columns: {len(hdr)}")
    print(f"  columns: {', '.join(hdr[:20])}")
    if N == 0:
        return

    def col(j):
        return [r[j] if j < len(r) else None for r in data]

    # error scan
    errs = sum(1 for r in data for v in r if v is not None and ERR.search(str(v)))
    if errs:
        print(f"  ⚠ ERROR cells (#REF! etc): {errs}")

    print("  --- columns ---")
    for j, h in enumerate(hdr):
        vals = col(j)
        nonblank = [v for v in vals if v is not None and str(v).strip() != ""]
        blanks = N - len(nonblank)
        distinct = len(set(str(v).strip() for v in nonblank))
        flags = []
        ws_trail = sum(1 for v in nonblank if str(v) != str(v).rstrip())
        ws_dbl = sum(1 for v in nonblank if "  " in str(v))
        if ws_trail: flags.append(f"trailing-space×{ws_trail}")
        if ws_dbl: flags.append(f"double-space×{ws_dbl}")
        line = f"    {h[:26]:<26} blank {100*blanks/N:5.1f}% | distinct {distinct}"
        if flags: line += "  [" + ", ".join(flags) + "]"
        print(line)
        # controlled vocabulary: low cardinality -> show top values
        if 0 < distinct <= 15 and len(nonblank) > 0:
            top = Counter(str(v).strip() for v in nonblank).most_common(8)
            print("        values: " + ", ".join(f"{k}×{c}" for k, c in top))

    # duplicate analysis on the key text column
    if key is None:
        # heuristic: the text column with the greatest average length
        best, blen = None, 0
        for j, h in enumerate(hdr):
            vals = [str(v) for v in col(j) if v is not None and str(v).strip() != ""]
            if not vals: continue
            avg = sum(len(v) for v in vals) / len(vals)
            if avg > blen and avg >= 6:
                best, blen = h, avg
        key = best
    if key and key in hdr:
        j = hdr.index(key)
        vals = [str(v) for v in col(j) if v is not None and str(v).strip() != ""]
        exact = [(k, c) for k, c in Counter(vals).items() if c > 1]
        groups = defaultdict(list)
        for v in vals:
            groups[norm(v)].append(v)
        nearclusters = [(k, vs) for k, vs in groups.items() if len(set(vs)) > 1]
        print(f"  --- duplicate scan on key column: {key!r} ---")
        print(f"    exact duplicates: {len(exact)} values across {sum(c for _, c in exact)} rows")
        for k, c in sorted(exact, key=lambda x: -x[1])[:5]:
            print(f"        {c}×  {k!r}")
        print(f"    near-dupe clusters (same after normalisation, differ raw): {len(nearclusters)}")
        for k, vs in sorted(nearclusters, key=lambda x: -len(x[1]))[:5]:
            print(f"        variants={sorted(set(vs))[:4]}")

def guess_shape(sheets):
    names = " ".join(s.upper() for s in sheets)
    hints = []
    if any(t in names for t in ["STOCK", "IN", "OUT", "ITEM DB", "GRN", "LEDGER"]):
        hints.append("structured-ledger (Excel) adapter")
    if any(t in names for t in ["GL", "JOURNAL", "FUND", "GRANT", "TRIAL BALANCE", "P&L", "QUICKBOOKS"]):
        hints.append("QuickBooks / accounting-export adapter")
    if any(t in names for t in ["PATIENT", "CLAIM", "POS", "SALES", "ENROL", "FEES", "LEARNER"]):
        hints.append("vertical-system-export adapter (Tier 2)")
    if len(sheets) == 1 and not hints:
        hints.append("formless-sheet adapter (single ad-hoc sheet) — verify header/structure")
    return hints or ["unrecognised — inspect manually, likely formless-sheet adapter"]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--key", default=None)
    args = ap.parse_args()
    path = args.file
    ext = os.path.splitext(path)[1].lower()
    print(f"FILE: {path}  ({os.path.getsize(path):,} bytes)")
    if ext in (".xlsx", ".xlsm"):
        sheets = load_xlsx(path)
    elif ext in (".csv",):
        sheets = load_csv(path)
    else:
        print(f"Unsupported extension {ext}. Convert to .xlsx or .csv first."); sys.exit(1)
    print(f"SHEETS ({len(sheets)}): {list(sheets.keys())}")
    print(f"SHAPE GUESS: {', '.join(guess_shape(list(sheets.keys())))}")
    for name, rows in sheets.items():
        if args.sheet and name != args.sheet:
            continue
        profile_sheet(name, rows, key=args.key)
    print(f"\n{'='*72}\nNext: map the detected columns to the OpsOS canonical fields, then run the")
    print("staging → validate → de-dup → confirm flow. See 08_Segment_Onboarding_Playbook.md.")

if __name__ == "__main__":
    main()
